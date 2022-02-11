from os import link
from bs4 import BeautifulSoup
from bs4.element import AttributeValueWithCharsetSubstitution
import requests
import lxml
import openpyxl

class website():
    def __init__(self, name, searchLink, link, tag, chosenClass, scrapingType):
        self.name = name
        self.searchLink = searchLink
        self.link = link
        self.tag = tag
        self.chosenClass = chosenClass
        self.scrapingStyle = scrapingType
    

    def scraping(self, argument):
        # "~°Ñ°~" will be placed where the search term should be, so we can replace it with the argument here
        self.searchLink.replace("~°Ñ°~", argument)
        r = requests.get(self.searchLink)
        soup = BeautifulSoup(r.content, "lxml")

        #These are the 2 lists we are returning
        Headlines = []
        Links = []
        
        #perhaps the select way is better for getting everything, I should try it.

        #Getting the headlines is easy
        for x in soup.find_all(f"{self.tag}", class_= f"{self.chosenClass}"):
            Headlines.append((x.get_text()).strip())

        #Getting the links not so much
        for y in soup.find_all(f"{self.tag}", class_= f"{self.chosenClass}", href = True):
            aquiredLink = y.get("href")

            #if the link is a string and its not only a part of the whole webpage link, we add it
            if ((aquiredLink.find(self.link) == 1) and (type(aquiredLink) == str)):
                Links.append(aquiredLink)
            
            #If it is a string but only a part of the real link, we add it plus the missing part
            elif ((aquiredLink.find(self.link) > 1) and (type(aquiredLink) == str)):
                Links.append(self.link + aquiredLink)
            
            #If the element didn't have any hrefs, we will search on its children
            elif (type(aquiredLink) != str):  
                children = str(y.findChildren())
                children = children.split('"')
                index = children.index(" href=")
                aquiredLink = children[index + 1]
                
                #Again, same conditions as before to check if the link is complete
                if ((aquiredLink.find(self.link) == 1) and (type(aquiredLink) == str)):
                        Links.append(aquiredLink)
                    
                elif ((aquiredLink.find(self.link) > 1) and (type(aquiredLink) == str)):
                        Links.append(self.link + aquiredLink)

        return(Headlines, Links)

    def writing(self, my_sheet, my_wb, HeadsAndLinks):

        #Checking if the first cell of a new column is empty, if it is it will start writing on it, if not it'll search for another
        m = 2
        while ((my_sheet.cell(row=1, column=m).value != None) or
        (my_sheet.cell(row= 1 , column = m).value != "")):
            m += 1

        titles = HeadsAndLinks[0]
        links = HeadsAndLinks[1]
        
        my_sheet.cell(row = 1, column = m).value = self.name

        n = 3
        for title in titles, links in links:
            my_sheet.cell(row = n, column = m).value = title
            my_sheet.cell(row = n+1, column = m).value = link
            n += 3
        
        my_wb.save("INVESTIGATOR.xlsx")

def initPyxl(self):
    #creating the xlsx file
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active

    #putting the number of the search results
    for n in range(1, 31):
        if (n % 3 == 0):
            my_sheet.cell(row=n, column=1).value = (n/3)

pubmed = website("Pubmed", "https://pubmed.ncbi.nlm.nih.gov/?term=~°Ñ°~", "https://pubmed.ncbi.nlm.nih.gov",
"a", "docsum-title", "basic")
scholar = website("Scholar", "https://scholar.google.com/scholar?hl=es&as_sdt=0%2C5&q=~°Ñ°~&btnG=", "https://scholar.google.com/" 
"h3", "gs_rt", "basic")
researchgate = website("ResearchGate","https://www.researchgate.net/search/publication?q=~°Ñ°~", "https://www.researchgate.net/", "div",
"nova-legacy-e-text nova-legacy-e-text--size-l nova-legacy-e-text--family-sans-serif nova-legacy-e-text--spacing-none nova-legacy-e-text--color-inherit nova-legacy-v-publication-item__title",
"basic")
elsevier = website("Elsevier", "https://www.elsevier.com/search-results?query=~°Ñ°~", "https://www.elsevier.com/", 
"h2", "search-result-title", "basic")
libgen = website("Libgen", "https://libgen.li/index.php?req=[~°Ñ°~&columns%5B%5D=t&columns%5B%5D=a&columns%5B%5D=s&columns%5B%5D=y&columns%5B%5D=p&columns%5B%5D=i&objects%5B%5D=f&objects%5B%5D=e&objects%5B%5D=s&objects%5B%5D=a&objects%5B%5D=p&objects%5B%5D=w&topics%5B%5D=l&topics%5B%5D=a&topics%5B%5D=m&topics%5B%5D=s&res=25"
"libgen.li", "", "", "libgen")

listOfPages = [pubmed, scholar, researchgate, elsevier, libgen]

