from bs4 import BeautifulSoup
from bs4.element import AttributeValueWithCharsetSubstitution
import requests
import lxml
import openpyxl

def startingSheet(webpages):
  my_wb = openpyxl.Workbook()
  my_sheet = my_wb.active
  my_sheet.title = "Investigación"

  for n in range(1, 11):
    if (n % 2 == 0):
      my_sheet.cell(row=n, column=1).value = n
      
   x = 2
   for page in webpages:
      my_sheet.cell(row=1, column=x).value = page
      x += 1
      
      
def writingData(list, y):
  x = 2
  for headline in list[0]:
     my_sheet.cell(row=x, column=y).value = headline
     x += 2
  
  x = 3
  for link in list[1]:
     my_sheet.cell(row=x, column=y).value = link
     x += 2

def scrapingWebsite(argument, tag, class):
  r = requests.get(argument)
  soup = BeautifulSoup(r.content, "lxml")
  Headlines = []
  Links = []

  #First of all: we get all the headlines
  for x in soup.find_all(f"{tag}", class_= f"{class}"):
     Headlines.append(x.get_text())

  for y in soup.find_all(f"{tag}", class_= f"{class}", href=True):
     Links.append(x[href])
  
  return([Headlines, Links])
      

def start(argument):
  # Page = [link, tag, class]
  
  pubmed = [f"https://pubmed.ncbi.nlm.nih.gov/?term={argument}", "a", "docsum-title"]
  scholar = [f"https://scholar.google.com/scholar?hl=es&as_sdt=0%2C5&q={argument}&btnG=", "h3", "gs_rt"]
  libgen = [f"https://libgen.li/index.php?req={argument}&columns%5B%5D=t&columns%5B%5D=a&columns%5B%5D=s&columns%5B%5D=y&columns%5B%5D=p&columns%5B%5D=i&objects%5B%5D=f&objects%5B%5D=e&objects%5B%5D=s&objects%5B%5D=a&objects%5B%5D=p&objects%5B%5D=w&topics%5B%5D=l&topics%5B%5D=c&topics%5B%5D=f&topics%5B%5D=a&topics%5B%5D=m&topics%5B%5D=r&topics%5B%5D=s&res=25",
  "", ""]
  sciencedir = [f"https://www.sciencedirect.com/search?qs={argument}", "a", "result-list-title-link"]
  researchgate = [f"https://www.researchgate.net/search/publication?q={argument}", "div",
  "nova-legacy-e-text nova-legacy-e-text--size-l nova-legacy-e-text--family-sans-serif nova-legacy-e-text--spacing-none nova-legacy-e-text--color-inherit nova-legacy-v-publication-item__title"]
  elsevier = [f"https://www.elsevier.com/search-results?query={argument}", "header", "search-result-header"]
  
  startingSheet(["PubMed", "Scholar", "LibGen" ,"Sciencedir","ResearchGate", "Elsevier"])
  
  position = 2
  for webpage in listOfWebpages:
    writingData(scrapingWebsite(webpage[0], webpage[1], webpage[2]), position)
    position += 1
  
start()
