from operator import truediv
from bs4 import BeautifulSoup
from bs4.element import AttributeValueWithCharsetSubstitution
import requests
import lxml
import openpyxl
import re

class website():
    def __init__(self, name, searchLink, link, tag, chosenClass, scrapingType):
        self.name = name
        self.searchLink = searchLink
        self.link = link
        self.tag = tag
        self.chosenClass = chosenClass
        self.scrapingStyle = scrapingType
    

    def scraping(self, argument):
        # "~°Ñ°~" is placed where the search term should be, so we can replace it with the argument here
        self.searchLink = self.searchLink.replace("~°Ñ°~", argument)

        r = requests.get(self.searchLink)
        soup = BeautifulSoup(r.content, "lxml")

        #These are the 2 lists we are returning
        Headlines = []
        Links = []
        
        if (self.scrapingStyle == "basic") :
            #Getting the headlines is easy
            for x in soup.find_all(f"{self.tag}", class_= f"{self.chosenClass}"):
                Headlines.append((x.get_text()).strip())
            

            #for y in soup.find_all(f"{self.tag}", class_= f"{self.chosenClass}", href=True):
            for y in soup.find_all(f'{self.tag}', class_= f"{self.chosenClass}", href=True):
                aquiredLink = y.find('href').text
                "aquiredLink = y.get('href')"
                if (True):
                    #if the link is a string and its not only a part of the whole webpage link, we add it
                    if ((aquiredLink.find("https://") == 1) or (aquiredLink.find("http://") == 1)):
                        Links.append(aquiredLink)

                    #If it is a string but only a part of the real link, we add it plus the missing part
                    else:
                        Links.append(self.link + aquiredLink)

            #If the element didn't have any hrefs, we will search on its children
            if (Links == []):  
                for y in soup.find_all(f"{self.tag}", class_= f"{self.chosenClass}", href = True):
                    children = str(y.findChildren())
                    children = children.split('"')
                    index = children.index(" href=")
                    aquiredLink = children[index + 1]

                    #Again, same conditions as before to check if the link is complete
                    if ((aquiredLink.find("https://") == 1) or (aquiredLink.find("http://") == 1)):
                        Links.append(aquiredLink)

                    else:
                        Links.append(self.link + aquiredLink)

                            
        #Libgen was not friendly to scraping due to not having classes, so i had to use the select method and iterate on a css selector.
        elif (self.scrapingStyle == "select"):
            for x in range(11):
                cssSelector = (self.tag).replace("~°Ñ°~", f"{x}")              
                elements = soup.select(cssSelector)
                for element in elements:
                    Headlines.append(element.get_text())
                    if ((element.get("href")).find(self.link) < 1): 
                        Links.append(self.link + element.get("href"))
                    else: 
                        Links.append(element.get("href"))

        print(f"{self.name} scrapping completed")
        return(Headlines, Links)


pubmed = website("Pubmed", "https://pubmed.ncbi.nlm.nih.gov/?term=~°Ñ°~", "https://pubmed.ncbi.nlm.nih.gov/",
"a", "docsum-title", "basic")

scholar = website( "Scholar", "https://scholar.google.com/scholar?hl=es&as_sdt=0%2C5&q=~°Ñ°~&btnG=", "",
"h3", "gs_rt", "basic")

researchgate = website("ResearchGate","https://www.researchgate.net/search/publication?q=~°Ñ°~", "https://www.researchgate.net/", "div",
"nova-legacy-e-text nova-legacy-e-text--size-l nova-legacy-e-text--family-sans-serif nova-legacy-e-text--spacing-none nova-legacy-e-text--color-inherit nova-legacy-v-publication-item__title",
"basic")

elsevier = website("Elsevier", "https://www.elsevier.com/search-results?query=~°Ñ°~", "https://www.elsevier.com/", 
"h2", "search-result-title", "basic")

libgen = website("Libgen", "https://libgen.is/scimag/?q=~°Ñ°~",
"https://libgen.is/", ".catalog > tbody:nth-child(2) > tr:nth-child(~°Ñ°~) > td:nth-child(2) > p:nth-child(1) > a:nth-child(1)", "", "select")

basesearch = website("Bielefeld Academic Search Engine", "https://www.base-search.net/Search/Results?lookfor=~°Ñ°~&name=&oaboost=1&newsearch=1&refid=dcbasen",
"https://www.base-search.net/" ,"a" ,"bold", "basic")

eric = website("ERIC", "https://eric.ed.gov/?q=~°Ñ°~", "https://eric.ed.gov/", "div", "r_t", "basic")

#doaj = website("Directory of Open Access journals and articles", 'https://www.doaj.org/search/journals?ref=homepage-box&source=%7B%22query%22%3A%7B%22query_string%22%3A%7B%22query%22%3A%22argentina%22%2C%22default_operator%22%3A%22AND%22%7D%7D%2C%22track_total_hits%22%3Atrue%7D',
#"https://www.doaj.org/", "li.card:nth-child(~°Ñ°~) > article:nth-child(1) > div:nth-child(1) > header:nth-child(1) > h3:nth-child(1) > a:nth-child(1)", "", "select")
#Unable to be scrapped



listOfPages = [scholar, researchgate, elsevier ,eric]

def writing(argument):
    #creating the xlsx file
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active

    #putting the number of the search results
    for n in range(1, 37):
        if (n % 3 == 0):
            my_sheet.cell(row=n, column=1).value = (n/3)

    rowCounter = 1
    columnCounter = 2

    for webpage in listOfPages:
        print(f"{webpage.name} started")

        my_sheet.cell(row= 1, column= columnCounter).value = webpage.name

        Returns = webpage.scraping(argument)
        Headlines = Returns[0]
        Links = Returns[1]

        rowCounter = 3
        for head in Headlines:
            my_sheet.cell(row = rowCounter, column = columnCounter).value = head
            rowCounter += 3

        rowCounter = 4
        for link in Links:
            my_sheet.cell(row = rowCounter, column = columnCounter).value = link
            rowCounter += 3

        columnCounter += 1

        print(f"{webpage.name} writen")

    my_wb.save("INVESTIGATOR.xlsx")
    print("Everything Finished!")


writing("zika")
