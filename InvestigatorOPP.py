from bs4 import BeautifulSoup
from bs4.element import AttributeValueWithCharsetSubstitution
import requests
import lxml
import openpyxl

from bs4 import BeautifulSoup
from bs4.element import AttributeValueWithCharsetSubstitution

##-------------------------------------------------------------------------------------------------------
class website():
    def __init__(self, name, searchLink, link, tag, chosenClass, scrapingType):
        self.name = name
        self.searchLink = searchLink
        self.link = link
        self.tag = tag
        self.chosenClass = chosenClass
        self.scrapingStyle = scrapingType
    
    #Returns = same argument or self.link + argument
    #If the argument does not start with http or https, it adds self.link to it.
    #Meant to check if the element obtained in scraping is actually a link or an ID.
    def checkIfHasLink(self, argument):
        if ( (argument.startswith("http:/") == True) or (argument.startswith("https:/") == True) ):
            return(argument)
        else: 
            return(self.link + argument)

    #Returns a list of lists, these lists contain 2 elements, a headline and a link.
    #List -> Lists -> Head and Link 
    def scraping(self, argument):
        # This variable is defined to be able to reset the value of searchLink at the end of the function
        originalSearchLink = self.searchLink

         # "~°Ñ°~" is placed where the search term should be, so we can replace it with the argument here
        self.searchLink = self.searchLink.replace("~°Ñ°~", argument)

        r = requests.get(self.searchLink)
        soup = BeautifulSoup(r.content, "lxml")

        contentO = []
        
        if (self.scrapingStyle == "basic") :
            #Neither the .get("href") and ["href"] methods were working, just kinda made my own
            allHtml = soup.find_all(f"{self.tag}", class_= f"{self.chosenClass}")

            for x in allHtml:

                head = (x.get_text()).strip()
                x = (str(x)).split('"')

                #After getting the headline, if for some reason getting a link is impossible, it breaks.
                try:
                    indexNumber = x.index(" href=") + 1
                    #Best way of getting the index of the link, due to href being one element away from it.
                except: break

                link = (self.checkIfHasLink(x[indexNumber]))

                #Appending a list with the headline and link.
                contentO.append([head, link])

                            
        #Another way to get everything, more efficient but its not always supported.
        elif (self.scrapingStyle == "select"):
            for x in range(10):
                cssSelector = (self.tag).replace("~°Ñ°~", f"{x}") #Setting up the selector. 
                selectedHtml = soup.select(cssSelector) #Selecting.
                if type(cssSelector) == str: #If it is a string, append a list with the headline and the link.
                    for x in selectedHtml:
                        contentO.append( [(x.get_text()).strip(), (self.checkIfHasLink(x.get("href")))] ) 
                        #Gets the text of the element and strips it, gets the link and uses the check if has link function on it.

        #Resets the value of self.searchLink
        self.searchLink = originalSearchLink

        return(contentO)
    
#-------------------------------------------------------------------------------------------------------
scholar = website( "Scholar", "https://scholar.google.com/scholar?hl=es&as_sdt=0%2C5&q=~°Ñ°~&btnG=", "",
"h3", "gs_rt", "basic")

researchgate = website("ResearchGate","https://www.researchgate.net/search/publication?q=~°Ñ°~", "https://www.researchgate.net/", "div",
"nova-legacy-e-text nova-legacy-e-text--size-l nova-legacy-e-text--family-sans-serif nova-legacy-e-text--spacing-none nova-legacy-e-text--color-inherit nova-legacy-v-publication-item__title",
"basic")

basesearch = website("Bielefeld-Academic-Search-Engine", "https://www.base-search.net/Search/Results?lookfor=~°Ñ°~&name=&oaboost=1&newsearch=1&refid=dcbasen",
"https://www.base-search.net/" ,"a" ,"bold", "basic")
    
pubmed = website("Pubmed", "https://pubmed.ncbi.nlm.nih.gov/?term=~°Ñ°~", "https://pubmed.ncbi.nlm.nih.gov/",
"article.full-docsum:nth-child(~°Ñ°~) > div:nth-child(2) > div:nth-child(1) > a:nth-child(1)", "", "select")

elsevier = website("Elsevier", "https://www.elsevier.com/search-results?query=~°Ñ°~", "https://www.elsevier.com/", 
"article.search-result:nth-child(~°Ñ°~) > header:nth-child(1) > h2:nth-child(1) > a", "", "select")

libgen = website("Libgen", "https://libgen.is/scimag/?q=~°Ñ°~",
"https://libgen.is", ".catalog > tbody:nth-child(2) > tr:nth-child(~°Ñ°~) > td:nth-child(2) > p:nth-child(1) > a:nth-child(1)", "", "select")

listOfPages = [scholar, researchgate, pubmed, elsevier, libgen, basesearch]

#-------------------------------------------------------------------------------------------------------
def progressBar(page):
    length = listOfPages.index(page)
    print(f"""
    ------------------Progress------------------
    :::::::::::::::::::::{round ((length * 16.7), 2)}%:::::::::::::::::
    --------------------------------------------
    """)
    
#-------------------------------------------------------------------------------------------------------
def writing(argument):
    #creating the xlsx file
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active

    #putting the number of the search results
    for n in range(1, 46):
        if (n % 3 == 0):
            my_sheet.cell(row=n, column=1).value = (n/3)

    rowCounter = 1
    columnCounter = 2

    for webpage in listOfPages:

        my_sheet.cell(row= 1, column= columnCounter).value = webpage.name

        Returns = webpage.scraping(argument)

        rowCounter = 3
        for x in Returns:
            my_sheet.cell(row = rowCounter, column = columnCounter).value = x[0]
            my_sheet.cell(row = (rowCounter + 1), column = columnCounter).value = x[1]
            rowCounter += 3

        columnCounter += 1
        progressBar(webpage)

    my_wb.save("INVESTIGATOR.xlsx")
    print(f"""
    ------------------Progreso------------------
    :::::::::::::::::::::100%:::::::::::::::::::::
    --------------------------------------------
    """)

#-------------------------------------------------------------------------------------------------------
def initialization():
    inp = input("Enter your search term: ")
    writing(inp)

initialization()
